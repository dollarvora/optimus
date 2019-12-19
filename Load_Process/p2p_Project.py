# Create loads for P2P

### Used SQL Table and actions:
#[Business_Planning].[dbo].[OTD_1_P2P_F_PARAMETERS_EMAIL_ADDRESS] : Get data, alter, delete, send
#[Business_Planning].[dbo].[OTD_1_P2P_F_PARAMETERS] : Get data, alter, delete, send
#[Business_Planning].[dbo].[OTD_1_P2P_F_PRIORITY_WITHOUT_INVENTORY] : Get data
#[Business_Planning].[dbo].[OTD_1_P2P_F_INVENTORY] : Get data
#OTD_1_P2P_D_INCLUDED_INVENTORY : Get data
#OTD_1_P2P_F_HISTORICAL : Send data

### Important path
# Folder where the excel workbook is saved
saveFolder='S:\Shared\Business_Planning\Personal\Lefebvre\S2\p2p\p2p_Project\p2p_Project\\' #Folder to save data

#from Import_Functions import *
#from LoadBuilder import LoadBuilder
from ParametersBox import *
from P2P_Functions import *
import pandas as pd
from openpyxl.styles import (PatternFill, colors, Alignment)
import Builder_tests.Read_And_Write as rw

# Parameters variables
dayTodayComplete= pd.datetime.now().replace(second=0, microsecond=0) #date to set in SQL for import date
dayToday= weekdays(0) #Date to display in report
printLoads = False #Print created loads
AutomaticRun = False # set to True to automate code
dest_filename = 'P2P_Summary_'+dayToday #Name of excel file with today's date

if not AutomaticRun:
    if not OpenParameters():  # If user cancel request
        sys.exit()

timeSinceLastCall('',False)
#####################################################################################################################
                                                ###Get SQL DATA
# ####################################################################################################################

# #If SQL queries crash
downloaded = False
numberOfTry = 0
while not downloaded and numberOfTry<3: # 3 trials, SQL Queries sometime crash for no reason
    numberOfTry+=1
    try:
        downloaded = True
####################

        ####################################################################################
                           ###  Email address Query
        ####################################################################################


        headerEmail='EMAIL_ADDRESS'
        SQLEmail = SQLConnection('CAVLSQLPD2\pbi2', 'Business_Planning', 'OTD_1_P2P_F_PARAMETERS_EMAIL_ADDRESS',headers=headerEmail)
        QueryEmail=""" SELECT distinct [EMAIL_ADDRESS]
         FROM [Business_Planning].[dbo].[OTD_1_P2P_F_PARAMETERS_EMAIL_ADDRESS]
         WHERE PROJECT = 'P2P'
        """
        #GET SQL DATA
        EmailList = [ item for sublist in SQLEmail.GetSQLData(QueryEmail) for item in sublist]#[ sublist for sublist in SQLEmail.GetSQLData(QueryEmail) ]  #[ item for sublist in SQLEmail.GetSQLData(SQLEmail) for item in sublist]



        ####################################################################################
                           ###  Parameters Query
        ####################################################################################

        headerParams='POINT_FROM,POINT_TO,LOAD_MIN,LOAD_MAX,DRYBOX,FLATBED,TRANSIT,PRIORITY_ORDER,SKIP'
        SQLParams = SQLConnection('CAVLSQLPD2\pbi2', 'Business_Planning', 'OTD_1_P2P_F_PARAMETERS',headers=headerParams)
        QueryParams=""" SELECT  [POINT_FROM]
              ,[POINT_TO]
              ,[LOAD_MIN]
              ,[LOAD_MAX]
              ,[DRYBOX]
              ,[FLATBED]
              ,[TRANSIT]
              ,[PRIORITY_ORDER]
              ,DAYS_TO
          FROM [Business_Planning].[dbo].[OTD_1_P2P_F_PARAMETERS]
          where IMPORT_DATE = (select max(IMPORT_DATE) from [Business_Planning].[dbo].[OTD_1_P2P_F_PARAMETERS])
          and SKIP = 0
          order by PRIORITY_ORDER
        """
        #GET SQL DATA
        DATAParams = [ Parameters(*sublist) for sublist in SQLParams.GetSQLData(QueryParams) ]  #[ item for sublist in SQLEmail.GetSQLData(SQLEmail) for item in sublist]


        ####################################################################################
                           ###  Parameters P2P ORDER (for excel sheet order)
        ####################################################################################

        QueryOrder=""" SELECT distinct  [POINT_FROM],[POINT_TO]
          FROM [Business_Planning].[dbo].[OTD_1_P2P_F_PARAMETERS]
          where IMPORT_DATE = (select max(IMPORT_DATE) from [Business_Planning].[dbo].[OTD_1_P2P_F_PARAMETERS])
          and SKIP = 0
          order by [POINT_FROM],[POINT_TO]
        """
        #GET SQL DATA
        P2POrder = SQLParams.GetSQLData(QueryOrder)

        ####################################################################################
                           ###  WishList Query
        ####################################################################################
        headerWishList = 'SALES_DOCUMENT_NUMBER,SALES_ITEM_NUMBER,SOLD_TO_NUMBER,POINT_FROM,SHIPPING_POINT,DIVISION,MATERIAL_NUMBER,Size_Dimensions,Lenght,Width,Height,stackability,Quantity,Priority_Rank,X_IF_MANDATORY'
        SQLWishList = SQLConnection('CAVLSQLPD2\pbi2', 'Business_Planning', 'OTD_2_PRIORITY_F_P2P',headers=headerWishList)
        QueryWishList= """SELECT  [SALES_DOCUMENT_NUMBER]
              ,[SALES_ITEM_NUMBER]
              ,[SOLD_TO_NUMBER]
              ,[POINT_FROM]
              ,[SHIPPING_POINT]
              ,[DIVISION]
              ,RTRIM([MATERIAL_NUMBER])
              ,RTRIM([Size_Dimensions])
              ,convert(int,CEILING([Length]))
              ,convert(int,CEILING([Width]))
              ,convert(int,CEILING([Height]))
              ,convert(int,[stackability])
              ,[Quantity]
              ,[Priority_Rank]
              ,[X_IF_MANDATORY]
              ,[OVERHANG]
          FROM [Business_Planning].[dbo].[OTD_1_P2P_F_PRIORITY_WITHOUT_INVENTORY]
          where [POINT_FROM] <>[SHIPPING_POINT] and Length<>0 and Width <> 0 and Height <> 0
          and concat (POINT_FROM,SHIPPING_POINT) in (select distinct concat([POINT_FROM],[POINT_TO]) from [Business_Planning].[dbo].[OTD_1_P2P_F_PARAMETERS]
          where IMPORT_DATE = (select max(IMPORT_DATE) from [Business_Planning].[dbo].[OTD_1_P2P_F_PARAMETERS])
          and SKIP = 0)
          order by [X_IF_MANDATORY] desc, Priority_Rank
        """

        OriginalDATAWishList = SQLWishList.GetSQLData(QueryWishList)
        DATAWishList=[WishListObj(*obj) for obj in OriginalDATAWishList]

        ####################################################################################
                           ###  INV Query
        ####################################################################################
        headerINV = ''#Not important here
        SQLINV = SQLConnection('CAVLSQLPD2\pbi2', 'Business_Planning', 'OTD_1_P2P_F_INVENTORY', headers=headerINV)
        # QueryINV = """SELECT  [SHIPPING_POINT]
        #       ,[MATERIAL_NUMBER]
        #       ,case when [QUANTITY] <0 then 0 else convert(int,QUANTITY) end as qty
        #       ,[AVAILABLE_DATE]
        #       ,[STATUS]
        #   FROM [Business_Planning].[dbo].[OTD_1_P2P_F_INVENTORY]
        #   where status = 'INVENTORY'
        #   order by SHIPPING_POINT, MATERIAL_NUMBER
        # """
        QueryINV = """select distinct SHIPPING_POINT
              ,[MATERIAL_NUMBER]
              ,case when sum(tempo.[QUANTITY]) <0 then 0 else convert(int,sum(tempo.QUANTITY)) end as [QUANTITY]
              , convert(DATE,GETDATE()) as [AVAILABLE_DATE]
              ,'INVENTORY' as [STATUS]
			  from(
		
		  SELECT  [SHIPPING_POINT]
              ,[MATERIAL_NUMBER]
              , [QUANTITY]
              ,[AVAILABLE_DATE]
              ,[STATUS]
          FROM [Business_Planning].[dbo].[OTD_1_P2P_F_INVENTORY]
          where status = 'INVENTORY'

		  union( select [SHIPPING_POINT]
              ,[MATERIAL_NUMBER]
              , [QUANTITY]
              ,GETDATE() as [AVAILABLE_DATE]
              ,'INVENTORY' as [STATUS]
			   FROM [Business_Planning].[dbo].[OTD_1_P2P_F_INVENTORY]
          where status in ('QA HOLD') and AVAILABLE_DATE between convert(DATE,GETDATE()-1) and GETDATE() --(select case when DATEPART(WEEKDAY,getdate()) = 6 then convert(DATE,GETDATE()+3) else convert(DATE,GETDATE() +1) end)
		 )) as tempo
		 group by SHIPPING_POINT
              ,[MATERIAL_NUMBER]
              ,  [AVAILABLE_DATE]
              , [STATUS]
	     order by SHIPPING_POINT, MATERIAL_NUMBER
                """


        OriginalDATAINV = SQLINV.GetSQLData(QueryINV)
        DATAINV = [INVObj(*obj) for obj in OriginalDATAINV]

        ####################################################################################
                           ###  QA HOLD Query
        ####################################################################################
        Query_QA = """ SELECT  [SHIPPING_POINT]
              ,[MATERIAL_NUMBER]
              , [QUANTITY]
              ,convert (DATE,[AVAILABLE_DATE]) as AVAILABLE_DATE
              ,[STATUS]
          FROM [Business_Planning].[dbo].[OTD_1_P2P_F_INVENTORY]
          where status = 'QA HOLD'
		  and AVAILABLE_DATE = (case when DATEPART(WEEKDAY,getdate()) = 6 then convert(DATE,GETDATE()+3) else convert(DATE,GETDATE() +1) end)
                        """
        OriginalDATA_QA = SQLINV.GetSQLData(Query_QA)
        #DATA_QA = []
        for obj in OriginalDATA_QA:
            DATAINV.append(INVObj(*obj)) #add QA HOLD with inv
                                        # we want the QA at the end of inv list, so the skus in QA will be the last to be chose



        ####################################################################################
        ###  Included Shipping_point
        ####################################################################################
        headerInclude = ''  # Not important here
        SQLInclude = SQLConnection('CAVLSQLPD2\pbi2', 'Business_Planning', 'OTD_1_P2P_D_INCLUDED_INVENTORY', headers=headerInclude)
        QueryInclude = """select SHIPPING_POINT_SOURCE ,SHIPPING_POINT_INCLUDE
                            from OTD_1_P2P_D_INCLUDED_INVENTORY
            """
        OriginalDATAInclude = SQLInclude.GetSQLData(QueryInclude)
        #DATAInclude = [] # defined in P2P_Functions
        global DATAInclude
        for obj in OriginalDATAInclude:
            DATAInclude.append(Included_Inv(*obj))


        ####################################################################################
        ###  Look if all point_from + shipping_point are in parameters
        ####################################################################################

        SQLMissing = SQLConnection('CAVLSQLPD2\pbi2', 'Business_Planning', 'OTD_1_P2P_F_PRIORITY', headers='')
        QueryMissing = """SELECT DISTINCT [POINT_FROM]
              ,[SHIPPING_POINT]
          FROM [Business_Planning].[dbo].[OTD_1_P2P_F_PRIORITY_WITHOUT_INVENTORY] 
          where CONCAT(POINT_FROM,SHIPPING_POINT) not in (
            select distinct CONCAT( [POINT_FROM],[POINT_TO])
            FROM [Business_Planning].[dbo].[OTD_1_P2P_F_PARAMETERS] where IMPORT_DATE = (select max(IMPORT_DATE) from [Business_Planning].[dbo].[OTD_1_P2P_F_PARAMETERS]) )
            and [POINT_FROM] <>[SHIPPING_POINT] 
        """
        DATAMissing = SQLMissing.GetSQLData(QueryMissing)
    except:
        downloaded=False
        print('SQL Query failed')

#If SQL Queries failed
if not downloaded:
    try:
        send_email(EmailList, dest_filename, 'SQL QUERIES FAILED')
    except:
        pass
    sys.exit()


timeSinceLastCall('Get SQL DATA')

# If there are missing P2P in parameters table
if not AutomaticRun:
    if DATAMissing != []:
        if not MissingP2PBox(DATAMissing):
            sys.exit()

timeSinceLastCall('',False)
#####################################################################################################################
                                                ### Excel Workbook declaration
#####################################################################################################################


wb = Workbook()
Warning_fill = PatternFill(fill_type="solid", start_color="FFFF00",end_color="FFFF00")
my_fill = PatternFill(fill_type="solid", start_color="a6a6a6",end_color="a6a6a6")

### Summary
wsSummary = wb.active
wsSummary.title = "SUMMARY"
wsSummary.append(['POINT_FROM','SHIPPING_POINT','NUMBER_OF_LOADS'])
wsSummary.column_dimensions['A'].width =13
wsSummary.column_dimensions['B'].width =16
wsSummary.column_dimensions['C'].width =19
for y in range(1, 4):
    wsSummary.cell(row=1, column=y).fill = my_fill


### Approved loads
wsApproved = wb.create_sheet("APPROVED")
wsApproved.append(['POINT_FROM','SHIPPING_POINT','LOAD_NUMBER','MATERIAL_NUMBER','QUANTITY','SIZE_DIMENSIONS',
                   'SALES_DOCUMENT_NUMBER','SALES_ITEM_NUMBER','SOLD_TO_NUMBER'])
for y in range(1, 10):
    wsApproved.cell(row=1, column=y).fill = my_fill
for letter in ['A','B','C','D','E','F','H','I']:
    wsApproved.column_dimensions[letter].width = 20
wsApproved.column_dimensions['G'].width = 27

### Unbooked
wsUnbooked = wb.create_sheet("UNBOOKED")
wsUnbooked.append(['POINT_FROM','MATERIAL_NUMBER','QUANTITY'])
for y in range(1, 4):
    wsUnbooked.cell(row=1, column=y).fill = my_fill
wsUnbooked.column_dimensions['A'].width = 13
wsUnbooked.column_dimensions['B'].width = 16
wsUnbooked.column_dimensions['C'].width = 10

### Booked unused
wsUnused = wb.create_sheet("BOOKED_UNUSED")
wsUnused.append(['POINT_FROM','MATERIAL_NUMBER','QUANTITY'])
for y in range(1, 4):
    wsUnused.cell(row=1, column=y).fill = my_fill
wsUnused.column_dimensions['A'].width =13
wsUnused.column_dimensions['B'].width =16
wsUnused.column_dimensions['C'].width =10

#####################################################################################################################
                                                ###Isolate perfect match
#####################################################################################################################
ListApprovedWish = []

# We iterate through wishlist to keep the priority order
for wish in DATAWishList:
    position = 0 #To not loop through all inv Data at each iteration
    for Iteration in range( wish.QUANTITY ):
        for It, inv in enumerate(DATAINV[position::]):
            if  EquivalentPlantFrom(inv.POINT,wish.POINT_FROM) and wish.MATERIAL_NUMBER==inv.MATERIAL_NUMBER and inv.QUANTITY>0: #wish.POINT_FROM==inv.POINT and
                if inv.Future: #QA of tomorrow, need to look if load is for today or later
                    InvToTake = False
                    for param in DATAParams:
                        if wish.POINT_FROM == param.POINT_FROM and wish.SHIPPING_POINT==param.POINT_TO and param.days_to>0:
                            InvToTake = True
                            break
                    if InvToTake:
                        inv.QUANTITY -= 1
                        wish.INV_ITEMS.append(inv)
                        position += It
                        break  # no need to look further
                else: # we give the inv to the wish item
                    inv.QUANTITY-=1
                    wish.INV_ITEMS.append(inv)
                    position+=It
                    break  # no need to look further


    if len(wish.INV_ITEMS) < wish.QUANTITY: #We give back taken inv if there is not enough units
        for invToGiveBack in wish.INV_ITEMS:
            invToGiveBack.QUANTITY+=1
        wish.INV_ITEMS = []
    else:
        ListApprovedWish.append(wish)


#####################################################################################################################
                                                ###Create Loads
#####################################################################################################################
for param in DATAParams: # for all P2P in parameters
    # Create data table to create loads
    tempoOnLoad=[]
    columnsHead=['QTY','MODEL','LENGTH','WIDTH','HEIGHT','NBR_PER_CRATE','STACK_LIMIT','OVERHANG']
    invData = []#pd.DataFrame([],columns=['QTY','MODEL','PLANT_TO','LENGTH','WIDTH','HEIGHT','NBR_PER_CRATE','STACK_LIMIT','OVERHANG'])
    for wish in ListApprovedWish:
        if wish.POINT_FROM==param.POINT_FROM  and wish.SHIPPING_POINT==param.POINT_TO and wish.QUANTITY>0:
            tempoOnLoad.append(wish)
            invData.append([1, wish.SIZE_DIMENSIONS,wish.LENGTH,wish.WIDTH,wish.HEIGHT,1,wish.STACKABILITY,wish.OVERHANG])#quantity is one, one box for each line
    models_data = pd.DataFrame(data = invData, columns=columnsHead)

    # Create loads
    result = param.LoadBuilder.build(models_data,param.LOADMAX, plot_load_done=printLoads)
    #Choose which wish to send in load based on selected crates and priority order
    for model in result:
        found = False
        for OnLoad in tempoOnLoad:
            if OnLoad.SIZE_DIMENSIONS == model and OnLoad.QUANTITY>0:
                OnLoad.QUANTITY =0
                found = True
                param.AssignedWish.append(OnLoad)
                break
        if not found: # If one returned crate doesn't match data
            print('Error in Perfect Match: impossible result.\n')

#####################################################################################################################
                                                ###Store unallocated units in inv pool
#####################################################################################################################
# If a wish item is not on a load, we give back his reserved inv
for wish in ListApprovedWish:
    if wish.QUANTITY>0:
        for inv in wish.INV_ITEMS:
            inv.QUANTITY+=1
        wish.INV_ITEMS=[]

#####################################################################################################################
                                                ### Try to Make the minimum number of loads for each P2P
#####################################################################################################################
for param in DATAParams:
    if len(param.LoadBuilder) < param.LOADMIN: #If the minimum isn't reached
        # create data table
        tempoOnLoad = []
        columnsHead = ['QTY', 'MODEL', 'LENGTH', 'WIDTH', 'HEIGHT', 'NBR_PER_CRATE', 'STACK_LIMIT', 'OVERHANG']
        invData = []
        for wish in DATAWishList:
            if wish.POINT_FROM==param.POINT_FROM and wish.SHIPPING_POINT == param.POINT_TO and wish.QUANTITY>0:
                position =0
                for Iteration in range(wish.QUANTITY):
                    for It, inv in enumerate(DATAINV[position::]):
                        if EquivalentPlantFrom(inv.POINT,wish.POINT_FROM) and inv.MATERIAL_NUMBER==wish.MATERIAL_NUMBER and inv.QUANTITY >0 and (not inv.Future or inv.Future and param.days_to>0):
                            inv.QUANTITY -= 1
                            wish.INV_ITEMS.append(inv)
                            position += It
                            break # no need to look further
                if len(wish.INV_ITEMS) < wish.QUANTITY:  # We give back taken inv
                    for invToGiveBack in wish.INV_ITEMS:
                        invToGiveBack.QUANTITY += 1
                    wish.INV_ITEMS=[]
                else:
                    tempoOnLoad.append(wish)
                    invData.append([1, wish.SIZE_DIMENSIONS, wish.LENGTH, wish.WIDTH, wish.HEIGHT, 1, wish.STACKABILITY, wish.OVERHANG])

        models_data = pd.DataFrame(data=invData, columns=columnsHead)
        # Create loads
        result = param.LoadBuilder.build(models_data, param.LOADMIN - len(param.LoadBuilder), plot_load_done=printLoads)
        # Choose wish items to put on loads
        for model in result:
            found = False
            for OnLoad in tempoOnLoad:
                if OnLoad.SIZE_DIMENSIONS == model and OnLoad.QUANTITY > 0:
                    OnLoad.QUANTITY = 0
                    found = True
                    param.AssignedWish.append(OnLoad)
                    break
            if not found:
                print('Error in Perfect Match: impossible result.\n')
        for wish in tempoOnLoad: #If it is not on loads, give back inv
            if wish.QUANTITY>0:
                for inv in wish.INV_ITEMS:
                    inv.QUANTITY+=1
                wish.INV_ITEMS=[]




#####################################################################################################################
                                                ### Try to Make the maximum number of loads for each P2P
#####################################################################################################################


for param in DATAParams:
    #if we haven't reached the max number of loads
    if len(param.LoadBuilder) < param.LOADMAX:
        # Create data table
        tempoOnLoad = []
        columnsHead = ['QTY', 'MODEL', 'LENGTH', 'WIDTH', 'HEIGHT', 'NBR_PER_CRATE', 'STACK_LIMIT', 'OVERHANG']
        invData = []
        for wish in DATAWishList:
            if wish.POINT_FROM==param.POINT_FROM and wish.SHIPPING_POINT == param.POINT_TO and wish.QUANTITY>0:
                position =0
                for Iteration in range(wish.QUANTITY):
                    for It, inv in enumerate(DATAINV[position::]) :
                        if EquivalentPlantFrom(inv.POINT,wish.POINT_FROM) and inv.MATERIAL_NUMBER==wish.MATERIAL_NUMBER and inv.QUANTITY >0 and (not inv.Future or inv.Future and param.days_to>0):
                            inv.QUANTITY -= 1
                            wish.INV_ITEMS.append(inv)
                            position += It
                            break # no need to look further
                if len(wish.INV_ITEMS) < wish.QUANTITY:  # We give back taken inv
                    for invToGiveBack in wish.INV_ITEMS:
                        invToGiveBack.QUANTITY += 1
                    wish.INV_ITEMS=[]
                else:
                    tempoOnLoad.append(wish)
                    invData.append([1, wish.SIZE_DIMENSIONS, wish.LENGTH, wish.WIDTH, wish.HEIGHT, 1, wish.STACKABILITY, wish.OVERHANG])

        models_data = pd.DataFrame(data=invData, columns=columnsHead)
        # Create loads
        result = param.LoadBuilder.build(models_data, param.LOADMAX, plot_load_done=printLoads)
        #choose wish items to put on loads
        for model in result:
            found = False
            for OnLoad in tempoOnLoad:
                if OnLoad.SIZE_DIMENSIONS == model and OnLoad.QUANTITY > 0:
                    OnLoad.QUANTITY = 0
                    found = True
                    param.AssignedWish.append(OnLoad)
                    break
            if not found:
                print('Error in Perfect Match: impossible result.\n')
        for wish in tempoOnLoad: #If it is not on loads, give back inv
            if wish.QUANTITY>0:
                for inv in wish.INV_ITEMS:
                    inv.QUANTITY+=1
                wish.INV_ITEMS=[]



#####################################################################################################################
                                                ###Test to save data
#####################################################################################################################
##to see created loads for each p2p
# print('\n\n\n\n\n\nresult\n\n\n\n\n\n')
# for param in DATAParams:
#     print(param.POINT_FROM,' _ ',param.POINT_TO)
#     print(len(param.LoadBuilder))
#     print(param.LoadBuilder.trailers_done)
#     print(param.LoadBuilder.get_loading_summary())

lineIndex=2 #To display a warning if number of loads is lower than parameters min
LoadIteration =0 #Number of each load

#SQL to send DATA
headersResult = 'POINT_FROM,SHIPPING_POINT,LOAD_NUMBER,MATERIAL_NUMBER,QUANTITY,SIZE_DIMENSIONS,SALES_DOCUMENT_NUMBER ,SALES_ITEM_NUMBER,SOLD_TO_NUMBER,IMPORT_DATE'
SQLResult = SQLConnection('CAVLSQLPD2\pbi2', 'Business_Planning', 'OTD_1_P2P_F_HISTORICAL',headers=headersResult)

for order in P2POrder: #To send data in excel workbook, order by : -point_from , -point_to
    for param in DATAParams:
        if param.POINT_FROM == order[0] and param.POINT_TO == order[1]:
            #section for summary worksheet
            wsSummary.append([param.POINT_FROM,param.POINT_TO,len(param.LoadBuilder)])
            if len(param.LoadBuilder) < param.LOADMIN:
                wsSummary.cell(row=lineIndex, column=3).fill = Warning_fill
            lineIndex+=1
            #Approved worksheet
            loads = param.LoadBuilder.get_loading_summary()
            if len(param.LoadBuilder)>0: #If some loads were created
                for line in range(len(loads)):
                    for Iteration in range(int(loads["QTY"][line])): #For every load of this kind
                        LoadIteration+=1
                        for column in loads.columns[4::]:# loop through size_dimentions
                            if loads[column][line] != '': # if there is some quantity
                                for QUANTITY in range(int(loads[column][line])): #For every unit of this crate on the load
                                    for wish in param.AssignedWish: # we associate a wish unit to this crate, then we save it
                                        if not wish.Finished and wish.SIZE_DIMENSIONS == column:
                                            wish.Finished=True
                                            valuesSQL = [(param.POINT_FROM, param.POINT_TO, LoadIteration, wish.MATERIAL_NUMBER, wish.ORIGINAL_QUANTITY,
                                                 column,wish.SALES_DOCUMENT_NUMBER,wish.SALES_ITEM_NUMBER,wish.SOLD_TO_NUMBER,dayTodayComplete)]

                                            wsApproved.append([param.POINT_FROM, param.POINT_TO, LoadIteration, wish.MATERIAL_NUMBER, wish.ORIGINAL_QUANTITY,
                                                 column,wish.SALES_DOCUMENT_NUMBER,wish.SALES_ITEM_NUMBER,wish.SOLD_TO_NUMBER])

                                            SQLResult.sendToSQL(valuesSQL)
                                            break

            break


#Assign left inv to wishList, to look for booked but unused
for wish in DATAWishList:
    if wish.QUANTITY == 0 and not wish.Finished:
        print('Error with wish: ', wish.lineToXlsx())
    if wish.QUANTITY>0:
        position = 0
        for Iteration in range( wish.QUANTITY ):
            for It, inv in enumerate(DATAINV[position::]):
                if  EquivalentPlantFrom(inv.POINT,wish.POINT_FROM) and wish.MATERIAL_NUMBER==inv.MATERIAL_NUMBER and inv.QUANTITY - inv.unused>0:
                    if inv.Future:  # QA of tomorrow, need to look if load is for today or later
                        InvToTake = False
                        for param in DATAParams:
                            if wish.POINT_FROM == param.POINT_FROM and wish.SHIPPING_POINT == param.POINT_TO and param.days_to > 0:
                                InvToTake = True
                                break
                        if InvToTake:
                            inv.unused += 1
                            position += It
                            break  # no need to look further
                    else:
                        inv.unused+=1
                        position+=It
                        break  # no need to look further


# send inv in wsUnused and wsUnbooked
for inv in DATAINV:
    if inv.unused>0:
        wsUnused.append([inv.POINT,inv.MATERIAL_NUMBER,inv.unused])
    if inv.QUANTITY - inv.unused >0:
        wsUnbooked.append([inv.POINT,inv.MATERIAL_NUMBER,inv.QUANTITY-inv.unused])




reference = [savexlsxFile(wb, saveFolder, dest_filename)]
send_email(EmailList, dest_filename, '', reference)
#os.system('start "excel" "'+str(reference[0])+'"') # To open excel workbook
